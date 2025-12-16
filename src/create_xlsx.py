import os
from openpyxl import Workbook, load_workbook
from typing import Dict, List, Union

def save_data_to_excel(data_dict: Dict[str, Union[List[str], List[float], List[List[float]]]], 
                      para_type: str, 
                      overwrite: bool = False) -> None:
    """
    将字典数据保存或追加到Excel文件，支持覆盖写入
    
    参数:
    data_dict -- 包含数据的字典，格式示例:
        {
            'sub': ['s1', 's2'],         # 受试者列表
            'method': ['bsl', 'ncs', ...], # 方法列表
            'para1': [数值列表]           # 参数值列表
        }
    para_type -- 参数类型名称 (如 'para1', 'para2'), 需与字典中的键名一致
    overwrite -- 是否覆盖现有文件 (True: 覆盖, False: 追加)
    """
    # 确定文件名
    filename = f"{para_type}_results.xlsx"
    
    # 检查字典中是否存在指定参数类型的键
    if para_type not in data_dict:
        raise ValueError(f"字典中缺少指定的参数类型: {para_type}")
    
    # 确保数据维度一致
    values = data_dict[para_type]
    values_flat = []
    
    # 如果数据是嵌套列表，展平它
    if isinstance(values, list) and all(isinstance(item, list) for item in values):
        values_flat = values
    else:
        # 将单个值列表转换为嵌套列表形式
        values_flat = [values] if not isinstance(values[0], list) else values
    
    # 检查数据维度是否匹配
    if len(data_dict['sub']) != len(values_flat):
        raise ValueError(f"sub数量({len(data_dict['sub'])})与参数值数量({len(values_flat)})不匹配")
    
    # 检查是否覆盖或文件不存在
    if overwrite or not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data" # type: ignore
        
        # 写入表头（方法名称）
        ws.cell(row=1, column=1, value=None) # type: ignore
        for col_idx, method in enumerate(data_dict['method'], start=2):
            ws.cell(row=1, column=col_idx, value=method) # type: ignore
            
        # 写入数据
        current_row = 2
        for idx, sub in enumerate(data_dict['sub']):
            ws.cell(row=current_row, column=1, value=sub) # type: ignore
            for col_idx, value in enumerate(values_flat[idx], start=2): # type: ignore
                ws.cell(row=current_row, column=col_idx, value=value) # type: ignore
            current_row += 1
    else:
        # 追加模式
        wb = load_workbook(filename)
        ws = wb.active
        
        # 检查标题行是否匹配
        header = [cell.value for cell in ws[1][1:]]  # 跳过A列 # type: ignore
        if header != data_dict['method']:
            raise ValueError("当前数据的方法列表与文件标题不匹配，无法追加")
        
        # 找到第一个空行
        current_row = ws.max_row + 1 # type: ignore
        
        # 写入新数据
        for idx, sub in enumerate(data_dict['sub']):
            ws.cell(row=current_row, column=1, value=sub) # type: ignore
            for col_idx, value in enumerate(values_flat[idx], start=2): # type: ignore
                ws.cell(row=current_row, column=col_idx, value=value) # type: ignore
            current_row += 1
    
    # 保存文件
    wb.save(filename)
    print(f"数据已{'覆盖' if overwrite else '追加'}写入到 {filename}")